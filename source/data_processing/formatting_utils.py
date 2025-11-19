"""Formatting helpers for Excel exports."""
from __future__ import annotations
from typing import Iterable
import re
import pandas as pd

__all__ = ["safe_filename", "write_excel_with_thousands"]

INVALID_PATTERN = re.compile(r"[^A-Za-z0-9._-]+")


def safe_filename(name: str, max_len: int = 120) -> str:
    cleaned = INVALID_PATTERN.sub("_", str(name)).strip("._") or "file"
    if len(cleaned) > max_len:
        cleaned = cleaned[:max_len]
    return cleaned


def write_excel_with_thousands(df: pd.DataFrame,
                               path: str,
                               thousand_cols: Iterable[str],
                               column_width: int = 12) -> None:
    """Write DataFrame to Excel applying thousand separator formatting for specified columns.

    Uses xlsxwriter if available; falls back to openpyxl for number format. If both formatting
    attempts fail, a plain export is written.
    """
    thousand_cols = list(thousand_cols)
    wrote_with_format = False
    try:
        with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False)
            workbook = writer.book
            worksheet = writer.sheets["Sheet1"]
            fmt = workbook.add_format({"num_format": "#,##0"})
            for col_idx, col_name in enumerate(df.columns):
                if col_name in thousand_cols:
                    worksheet.set_column(col_idx, col_idx, column_width, fmt)
        wrote_with_format = True
    except Exception as e:
        print(f"xlsxwriter formatting failed for {path}: {e}. Falling back to openpyxl.")

    if wrote_with_format:
        return

    try:
        df.to_excel(path, index=False)
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        wb = load_workbook(path)
        ws = wb.active
        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_name in thousand_cols:
                col_letter = get_column_letter(col_idx)
                for cell in ws[col_letter]:
                    if cell.row == 1:
                        continue
                    if cell.value is None:
                        cell.value = 0
                    cell.number_format = '#,##0'
        wb.save(path)
    except Exception as e2:
        print(f"openpyxl formatting fallback failed for {path}: {e2}. Writing plain Excel.")
        df.to_excel(path, index=False)
